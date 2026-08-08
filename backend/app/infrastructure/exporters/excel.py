from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.domain.reports.configuration import ReportConfiguration
from app.infrastructure.exporters.common import format_value, protect_formula, visible_columns


class ExcelReportExporter:
    format = "xlsx"
    extension = "xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def export(
        self, path: Path, configuration: ReportConfiguration, rows: Iterable[dict[str, Any]]
    ) -> int:
        workbook = Workbook(write_only=False)
        sheet = workbook.active
        sheet.title = "Reporte"
        columns = visible_columns(configuration)
        sheet.append([item.label for item in columns])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        count = 0
        for row in rows:
            values: list[Any] = []
            for column in columns:
                value = row.get(column.source_key)
                if isinstance(value, (datetime, date, int, float, bool)) and column.format.type in {
                    "automatic",
                    "integer",
                    "decimal",
                    "date",
                    "datetime",
                    "boolean",
                }:
                    rendered = value
                else:
                    rendered = format_value(value, column)
                values.append(protect_formula(rendered))
            sheet.append(values)
            count += 1
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(columns, 1):
            sheet.column_dimensions[get_column_letter(index)].width = min(
                80, max(10, (column.width or 120) / 8)
            )
        workbook.save(path)
        return count
